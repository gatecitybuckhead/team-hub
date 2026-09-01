#!/usr/bin/env python3
"""Parse GCB 'Sunday DEBRIEF Form (Responses)' into tidy JSON."""
import json, re, sys, datetime, os, pathlib
from openpyxl import load_workbook
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
# Shared with tools/extract_debrief.py (the browser-free Drive path) so the two
# ingest paths can never produce different records for the same form row.
from debrief_common import (NAMES, ELEMENTS, COMMENT_COLS, MSG,
                            rating, ten, name_canon, sunday_for, avg)

ROOT = pathlib.Path(__file__).resolve().parent.parent
# Usage: python3 parse_debrief.py <Sunday DEBRIEF Form (Responses).xlsx> [out.json]
SRC = sys.argv[1] if len(sys.argv) > 1 else str(ROOT/'Sunday DEBRIEF Form (Responses).xlsx')
OUT = sys.argv[2] if len(sys.argv) > 2 else str(ROOT/'data'/'debrief.json')

wb = load_workbook(SRC, data_only=True)
ws = wb['Form Responses 1']
responses=[]
for r in range(2, ws.max_row+1):
    ts = ws.cell(r,1).value
    if not isinstance(ts, datetime.datetime): continue
    d = ts.date()
    sunday = sunday_for(d)
    ratings={}
    for el,(c1,c2) in ELEMENTS.items():
        v = rating(ws.cell(r,c1).value)
        if v is None and c2: v = rating(ws.cell(r,c2).value)
        if v is not None: ratings[el]=v
    msg={k:ten(ws.cell(r,c).value) for k,c in MSG.items()}
    msg['call_to_action'] = ten(ws.cell(r,31).value) or ten(ws.cell(r,33).value)
    msg={k:v for k,v in msg.items() if v is not None}
    comments={}
    for c,tag in COMMENT_COLS.items():
        v=str(ws.cell(r,c).value or '').strip()
        if v and v.upper() not in ('N/A','NA','NONE','N/A.','-'): comments[tag]=v
    responses.append({'sunday':sunday.isoformat(),'submitted':d.isoformat(),
        'name':name_canon(ws.cell(r,2).value),
        'overall':ten(ws.cell(r,3).value),'ratings':ratings,'message':msg,'comments':comments})

responses.sort(key=lambda x:(x['sunday'],x['name']))
# weekly aggregates
weeks={}
for resp in responses:
    w=weeks.setdefault(resp['sunday'],{'sunday':resp['sunday'],'n':0,'overall':[],
        'elements':{},'message':{},'comments':[],'names':[]})
    w['n']+=1; w['names'].append(resp['name'])
    if resp['overall']: w['overall'].append(resp['overall'])
    for k,v in resp['ratings'].items(): w['elements'].setdefault(k,[]).append(v)
    for k,v in resp['message'].items(): w['message'].setdefault(k,[]).append(v)
    for tag,t in resp['comments'].items(): w['comments'].append({'by':resp['name'],'tag':tag,'text':t})
weekly=[]
for k in sorted(weeks):
    w=weeks[k]
    weekly.append({'sunday':k,'n':w['n'],'names':sorted(set(w['names'])),
        'overall_avg':avg(w['overall']),
        'elements':{e:avg(v) for e,v in w['elements'].items()},
        'message':{m:avg(v) for m,v in w['message'].items()},
        'comments':w['comments']})
json.dump({'generated':datetime.date.today().isoformat(),
           'responses':responses,'weekly':weekly}, open(OUT,'w'), indent=1)
print('responses:',len(responses),'weeks:',len(weekly))
print('range:',weekly[0]['sunday'],'->',weekly[-1]['sunday'])
print('latest week:',json.dumps(weekly[-1],indent=1)[:600])
