#!/usr/bin/env python3
"""Parse GCB 'DASHBOARD METRICS - 5 Behaviors' quarterly tabs into tidy JSON."""
import json, os, re, sys, datetime
from openpyxl import load_workbook
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from metrics_common import clean, slug, ALIASES, canon, GIVING_KEYS

# Full-history rebuild only. Needs an xlsx export of the sheet, because the Drive
# connector truncates it (~146K chars). Pass paths on the command line:
#   python3 tools/parse_metrics.py "<in.xlsx>" [out.json]
# For a normal weekly increment use tools/add_metrics_week.py instead.
_HERE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SRC = sys.argv[1] if len(sys.argv) > 1 else os.path.join(
    os.path.expanduser('~'), 'Downloads', 'DASHBOARD METRICS - 5 Behaviors.xlsx')
OUT = sys.argv[2] if len(sys.argv) > 2 else os.path.join(_HERE, 'data', 'metrics.json')

TABS = [('Q1 2025',2025),('Q2 2025',2025),('Q3 2025',2025),('Q4 2025',2025),
        ('Q1 2026',2026),('Q2 2026',2026),('Q3 2026',2026)]

SECTIONS = {'SCORE','BOARD','PRAY','SERVE','GIVE','FORM','REACH','DIGITAL','GENERAL','Quarterly'}

# clean() / slug() / ALIASES / canon() now live in tools/metrics_common.py so the
# weekly-increment path (add_metrics_week.py) can't drift from this full rebuild.

# Manual series corrections — the sheet's "Series" row lags behind the actual
# preaching series. date (ISO) -> (series, special-or-None). Applied after parse.
# Add a new entry each time the series changes until the sheet catches up.
SERIES_OVERRIDES = {
    '2026-07-19': ('Journey Through James', 'Week #1'),
}

wb = load_workbook(SRC, data_only=True)
weeks = {}          # date -> {"date":..,"series":..,"special":..,"values":{}}
catalog = {}        # key -> {"label","section","owner"}

for tabname, year in TABS:
    ws = wb[tabname]
    # find header row: col A or B contains 'Data Set/Sunday'
    hdr = None
    for r in range(1,12):
        for c in (1,2):
            if 'Data Set/Sunday' in str(ws.cell(r,c).value or ''): hdr = r
        if hdr: break
    if not hdr:
        print('SKIP', tabname); continue
    series_row = None
    for rr in range(max(1,hdr-2), hdr):
        if 'Series' in str(ws.cell(rr,1).value or ''): series_row = rr
    special_row = hdr+1 if 'Special' in str(ws.cell(hdr+1,1).value or '') else None
    # map columns: sunday columns and their preceding data-set columns
    suncols = {}   # col -> date
    datacols = {}  # sunday col -> dataset col
    lastdata = None
    for c in range(2, ws.max_column+1):
        h = str(ws.cell(hdr,c).value or '').strip()
        if h.startswith('Data Set'):
            lastdata = c
        m = re.match(r'(?:Sunday\s+)?(\d{1,2})/(\d{1,2})\s*$', h)
        if m:
            mo,da = int(m.group(1)), int(m.group(2))
            y = year
            # Q1 tabs can reference a late-Dec Sunday of prior year
            if tabname.startswith('Q1') and mo==12: y = year-1
            try: d = datetime.date(y,mo,da)
            except ValueError: continue
            suncols[c] = d.isoformat()
            datacols[c] = lastdata
            lastdata = None
    # carry-forward series names across merged/blank cells
    for c,dt in sorted(suncols.items()):
        wk = weeks.setdefault(dt, {'date':dt,'series':None,'special':None,'values':{}})
        if series_row:
            sv=None
            for cc in range(c,1,-1):
                x = ws.cell(series_row,cc).value
                if x and str(x).strip(): sv=str(x).strip(); break
            if sv and not wk['series']: wk['series']=sv
        if special_row:
            x = ws.cell(special_row,c).value or (datacols[c] and ws.cell(special_row,datacols[c]).value)
            if x and str(x).strip(): wk['special']=str(x).strip()
    # metrics rows
    section='SCORE'; owner=None
    for r in range(hdr+1, ws.max_row+1):
        a = str(ws.cell(r,1).value or '').strip()
        label = str(ws.cell(r,2).value or '').strip()
        if a:
            base = a.rstrip(':').strip()
            if base in SECTIONS: section = base; owner=None
            m = re.match(r'^\((.+)\)$', base)
            if m: owner = m.group(1).strip()
        if not label or label=='Special Service:': continue
        if not section: continue
        key = canon(slug(label))
        if key not in catalog:
            catalog[key] = {'label':label,'section':section,'owner':owner}
        elif owner and not catalog[key]['owner']:
            catalog[key]['owner']=owner
        for c,dt in suncols.items():
            sun_v = clean(ws.cell(r,c).value)
            ds_v = clean(ws.cell(r,datacols[c]).value) if datacols[c] else None
            if key in GIVING_KEYS and (sun_v is not None or ds_v is not None):
                # keep both cells — the week is Mon–Sat + Sunday (see GIVING_KEYS)
                weeks[dt].setdefault('giving_cols', {})[key] = {'ds': ds_v, 'sun': sun_v}
            v = sun_v if sun_v is not None else ds_v
            if v is not None:
                weeks[dt]['values'][key] = v

for dt,(sv,sp) in SERIES_OVERRIDES.items():
    if dt in weeks:
        weeks[dt]['series'] = sv
        if sp is not None: weeks[dt]['special'] = sp

today = datetime.date.today().isoformat()
kept = {k:v for k,v in weeks.items() if k <= today}   # drop stale template columns dated in the future
out = {'generated': today,
       'catalog': catalog,
       'weeks': [kept[k] for k in sorted(kept)]}
json.dump(out, open(OUT,'w'), indent=1)
ws_with_data = [w for w in out['weeks'] if w['values']]
print('weeks:', len(out['weeks']), 'with data:', len(ws_with_data))
print('range:', out['weeks'][0]['date'], '->', out['weeks'][-1]['date'])
print('metrics:', len(catalog))
